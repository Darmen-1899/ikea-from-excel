from selenium.webdriver import ActionChains
from selenium import webdriver
import openpyxl

driver = webdriver.Chrome(executable_path='chromedriver.exe')


wb_write = openpyxl.load_workbook('готовый лист.xlsx')
wb_write.create_sheet('Sheet1')
worksheet = wb_write['Sheet1']

wb_read = openpyxl.load_workbook('Икеа номенклатура (4000 товаров) (1).xltx')
active_wb_read = wb_read.active
max_rows_in_read_wb = active_wb_read.max_row


index = 1

for i in range(2,max_rows_in_read_wb + 1):
    product_status = ' '
    product_price = ' '
    product_name = active_wb_read.cell(row=i, column=1).value
    product_articul = active_wb_read.cell(row=i, column=2).value
    product_url = active_wb_read.cell(row=i, column=4).value
    driver.get(product_url)

    #Cookies
    try:
        button = driver.find_element_by_class_name(u'js-cookie-info__accept-button')
        driver.implicitly_wait(1)
        ActionChains(driver).move_to_element(button).click(button).perform()
    except Exception:
        print("No cookies")

    #Price
    try:
        product_price = driver.find_element_by_xpath('//*[@id="content"]/div/div/div/div[2]/div[3]/div/div[1]/div/div[2]/div').text
    except Exception:
        print('No price')
    try:
        product_price = driver.find_elements_by_class_name('range-revamp-pip-price-package__main-price').text
    except Exception:
        print("No price")

    #offline shop
    try:
        driver.find_element_by_xpath('//*[@id="content"]/div/div/div/div[2]/div[3]/div/div[4]/div[2]/div').click()
    except Exception:
        print('Another link')
    try:
        driver.find_element_by_link_text('Проверка наличия в офлайн-магазине').click()
    except Exception:
        print('Another link')
    try:
        driver.find_element_by_class_name(
            'range-revamp-stockcheck__available-for-delivery-link range-revamp-link').click()
    except Exception:
        print('Another link')


    try:
        button1 = driver.find_element_by_xpath('//*[@id="change-store-input"]')
        ActionChains(driver).move_to_element(button1).send_keys('омск').perform()
        if driver.find_element_by_class_name('range-revamp-stockcheck__store-text').text == "В наличии":
            product_status = "В наличии"
        elif driver.find_element_by_class_name('range-revamp-stockcheck__store-text').text == "Заканчивается":
            product_status = "Заканчивается"
        elif driver.find_element_by_class_name('range-revamp-stockcheck__store-text').text == "Почти закончился":
            product_status = "Почти закончился"
        else:
            product_status = "Нет в наличии"
    except Exception:
        product_status = 'No search'
    worksheet['A' + str(index)] = product_name
    worksheet['B' + str(index)] = product_articul
    worksheet['C' + str(index)] = product_url
    worksheet['D' + str(index)] = product_price
    worksheet['E' + str(index)] = product_status
    index = index + 1
    wb_write.save('готовый лист.xlsx')
    print(product_name + ' ' + product_articul + ' ' + product_url + ' ' + product_price)
    print(index)

wb_write.save('готовый лист.xlsx')
driver.close()


